"""Excel export service using openpyxl."""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.database import get_db
from app.parser.codes import lookup_status, lookup_group, lookup_carc


def export_file_to_excel(file_id: int) -> bytes:
    """Export a complete file with all data to an Excel workbook."""
    db = get_db()
    try:
        # Get file info
        file_row = db.execute("SELECT * FROM edi_files WHERE id = ?", (file_id,)).fetchone()
        if not file_row:
            raise ValueError("File not found")
        file_info = dict(file_row)

        wb = Workbook()
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
        thin_border = Border(
            bottom=Side(style="thin", color="DEE2E6"),
        )

        def style_header(ws):
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            ws.freeze_panes = "A2"

        def auto_width(ws):
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

        # --- Summary Sheet ---
        ws = wb.active
        ws.title = "Summary"
        summary_data = [
            ("Field", "Value"),
            ("Filename", file_info.get("filename", "")),
            ("Uploaded", file_info.get("uploaded_at", "")),
            ("Payer", file_info.get("payer_name", "")),
            ("Payer ID", file_info.get("payer_id", "")),
            ("Payee", file_info.get("payee_name", "")),
            ("Payee ID", file_info.get("payee_id", "")),
            ("Payment Amount", file_info.get("bpr_amount", 0)),
            ("Payment Date", file_info.get("bpr_payment_date", "")),
            ("Payment Method", file_info.get("bpr_payment_method", "")),
            ("TRN Reference", file_info.get("trn_reference", "")),
            ("Contact", file_info.get("contact_name", "")),
            ("Phone", file_info.get("contact_phone", "")),
        ]
        for row in summary_data:
            ws.append(row)
        style_header(ws)
        auto_width(ws)

        # --- Claims Sheet ---
        ws_claims = wb.create_sheet("Claims")
        claims_headers = [
            "Claim ID", "Status Code", "Status", "Patient", "Patient ID",
            "Provider", "Provider ID", "Charges", "Payment", "Adjustments",
            "Plan Code", "DRG", "Date Start", "Date End", "Date Received",
        ]
        ws_claims.append(claims_headers)

        claims = db.execute(
            "SELECT * FROM claims WHERE file_id = ? ORDER BY id", (file_id,)
        ).fetchall()
        for c in claims:
            d = dict(c)
            ws_claims.append([
                d.get("clp_claim_id", ""),
                d.get("clp_status_code", ""),
                lookup_status(d.get("clp_status_code", "")),
                d.get("patient_name", ""),
                d.get("patient_id", ""),
                d.get("rendering_provider_name", ""),
                d.get("rendering_provider_id", ""),
                d.get("clp_total_charge", 0),
                d.get("clp_total_payment", 0),
                d.get("total_adjustments", 0),
                d.get("clp_plan_code", ""),
                d.get("clp_drg_code", ""),
                d.get("claim_date_start", ""),
                d.get("claim_date_end", ""),
                d.get("claim_received_date", ""),
            ])
        style_header(ws_claims)
        auto_width(ws_claims)

        # --- Service Lines Sheet ---
        ws_svc = wb.create_sheet("Service Lines")
        svc_headers = [
            "Claim ID", "Procedure", "Modifiers", "Revenue Code",
            "Charge", "Payment", "Units", "Date Start", "Date End", "Control #",
        ]
        ws_svc.append(svc_headers)

        for c in claims:
            cd = dict(c)
            svcs = db.execute(
                "SELECT * FROM service_lines WHERE claim_id = ? ORDER BY id", (cd["id"],)
            ).fetchall()
            for s in svcs:
                sd = dict(s)
                ws_svc.append([
                    cd.get("clp_claim_id", ""),
                    sd.get("procedure_code", ""),
                    sd.get("procedure_modifiers", ""),
                    sd.get("revenue_code", ""),
                    sd.get("charge_amount", 0),
                    sd.get("payment_amount", 0),
                    sd.get("units", 0),
                    sd.get("date_start", ""),
                    sd.get("date_end", ""),
                    sd.get("control_number", ""),
                ])
        style_header(ws_svc)
        auto_width(ws_svc)

        # --- Adjustments Sheet ---
        ws_adj = wb.create_sheet("Adjustments")
        adj_headers = [
            "Claim ID", "Level", "Group Code", "Group", "Reason Code", "Reason", "Amount", "Quantity",
        ]
        ws_adj.append(adj_headers)

        for c in claims:
            cd = dict(c)
            # Claim-level adjustments
            cadjs = db.execute(
                "SELECT * FROM claim_adjustments WHERE claim_id = ?", (cd["id"],)
            ).fetchall()
            for a in cadjs:
                ad = dict(a)
                ws_adj.append([
                    cd.get("clp_claim_id", ""),
                    "Claim",
                    ad.get("group_code", ""),
                    lookup_group(ad.get("group_code", "")),
                    ad.get("reason_code", ""),
                    lookup_carc(ad.get("reason_code", "")),
                    ad.get("amount", 0),
                    ad.get("quantity", 0),
                ])
            # Service-level adjustments
            svcs = db.execute(
                "SELECT * FROM service_lines WHERE claim_id = ?", (cd["id"],)
            ).fetchall()
            for s in svcs:
                sd = dict(s)
                sadjs = db.execute(
                    "SELECT * FROM service_adjustments WHERE service_line_id = ?", (sd["id"],)
                ).fetchall()
                for sa in sadjs:
                    sad = dict(sa)
                    ws_adj.append([
                        cd.get("clp_claim_id", ""),
                        f"Service ({sd.get('procedure_code', '')})",
                        sad.get("group_code", ""),
                        lookup_group(sad.get("group_code", "")),
                        sad.get("reason_code", ""),
                        lookup_carc(sad.get("reason_code", "")),
                        sad.get("amount", 0),
                        sad.get("quantity", 0),
                    ])
        style_header(ws_adj)
        auto_width(ws_adj)

        # --- Provider Adjustments Sheet ---
        ws_prov = wb.create_sheet("Provider Adjustments")
        prov_headers = ["Provider ID", "Fiscal Period", "Reason Code", "Amount"]
        ws_prov.append(prov_headers)

        provs = db.execute(
            "SELECT * FROM provider_adjustments WHERE file_id = ?", (file_id,)
        ).fetchall()
        for p in provs:
            pd_row = dict(p)
            ws_prov.append([
                pd_row.get("provider_id", ""),
                pd_row.get("fiscal_period_end", ""),
                pd_row.get("reason_code", ""),
                pd_row.get("amount", 0),
            ])
        style_header(ws_prov)
        auto_width(ws_prov)

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
    finally:
        db.close()
